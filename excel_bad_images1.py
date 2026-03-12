import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------------- FILE PATHS ----------------
rating_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_rating_202602111506.csv"
details_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_details_202602111506.csv"

bad_images_folder = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/rating_image_1/bad_images"

output_excel = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/bad_images_report1.xlsx"

# ---------------- LOAD CSV ----------------
rating_df = pd.read_csv(rating_csv, low_memory=False)
details_df = pd.read_csv(details_csv, low_memory=False)

# Clean column names
rating_df.columns = rating_df.columns.str.strip()
details_df.columns = details_df.columns.str.strip()

# Normalize rating text
rating_df["rating_value"] = rating_df["rating_value"].astype(str).str.lower().str.strip()
rating_df["rating_value_ai"] = rating_df["rating_value_ai"].astype(str).str.lower().str.strip()

# ---------------- MERGE ----------------
merged_df = pd.merge(
    rating_df,
    details_df,
    on="inspection_id",
    how="inner",
    suffixes=("_rating", "_details")
)

# ---------------- FILTER ONLY POOR + BAD ----------------
bad_df = merged_df[
    (merged_df["rating_value"] == "poor") &
    (merged_df["rating_value_ai"] == "bad")
]

# Keep only downloaded images
downloaded_images = os.listdir(bad_images_folder)
bad_df = bad_df[bad_df["image_id"].isin(downloaded_images)]

# ---------------- LIMIT TO 100 IMAGES ----------------
bad_df = bad_df.head(100)

# ---------------- CREATE EXCEL ----------------
wb = Workbook()
ws = wb.active
ws.title = "Bad Images"

headers = [
    "Image",
    "location",
    "rating_value",
    "rating_value_ai",
    "rating_my",
    "prs_coach_no",
    "staff_id",
    "updated_by",
    "image_id",
    "inspection_id"
]

ws.append(headers)

row_num = 2

for index, row in bad_df.iterrows():

    image_path = os.path.join(bad_images_folder, row["image_id"])

    # Insert text data
    ws.cell(row=row_num, column=2, value=row.get("location", ""))
    ws.cell(row=row_num, column=3, value=row.get("rating_value", ""))
    ws.cell(row=row_num, column=4, value=row.get("rating_value_ai", ""))
    ws.cell(row=row_num, column=5, value="")  # rating_my (editable)
    ws.cell(row=row_num, column=6, value=row.get("prs_coach_no", ""))
    ws.cell(row=row_num, column=7, value=row.get("staff_id", ""))
    ws.cell(row=row_num, column=8, value=row.get("updated_by_details", ""))
    ws.cell(row=row_num, column=9, value=row.get("image_id", ""))
    ws.cell(row=row_num, column=10, value=row.get("inspection_id", ""))

    # Insert BIG image
    if os.path.exists(image_path):
        img = XLImage(image_path)
        img.width = 400
        img.height = 400
        ws.add_image(img, f"A{row_num}")
        ws.row_dimensions[row_num].height = 310

    row_num += 1

# ---------------- FORMAT COLUMNS ----------------
column_widths = [57, 12, 12, 12, 12, 12, 12, 18, 18, 18]

for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Save file
wb.save(output_excel)

print("✅ Excel created successfully with 100 big images (no hyperlink needed)")
print("Saved at:", output_excel)
