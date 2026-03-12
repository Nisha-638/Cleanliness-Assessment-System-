# import pandas as pd
# import os
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image as XLImage
# from openpyxl.utils import get_column_letter

# # ---------------- FILE PATHS ----------------
# rating_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_rating_202602111506.csv"
# details_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_details_202602111506.csv"

# bad_images_folder = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/rating_image_1/bad_images"

# output_excel = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/bad_images_report.xlsx"

# # ---------------- LOAD CSV ----------------
# rating_df = pd.read_csv(rating_csv, low_memory=False)
# details_df = pd.read_csv(details_csv, low_memory=False)

# rating_df.columns = rating_df.columns.str.strip()
# details_df.columns = details_df.columns.str.strip()

# rating_df["rating_value"] = rating_df["rating_value"].astype(str).str.lower().str.strip()
# rating_df["rating_value_ai"] = rating_df["rating_value_ai"].astype(str).str.lower().str.strip()

# # ---------------- MERGE ----------------
# merged_df = pd.merge(
#     rating_df,
#     details_df,
#     on="inspection_id",
#     how="inner",
#     suffixes=("_rating", "_details")
# )

# # ---------------- FILTER BAD ----------------
# bad_df = merged_df[
#     (merged_df["rating_value"] == "poor") &
#     (merged_df["rating_value_ai"] == "bad")
# ]

# downloaded_images = os.listdir(bad_images_folder)
# bad_df = bad_df[bad_df["image_id"].isin(downloaded_images)]

# # ---------------- CREATE EXCEL ----------------
# # ---------------- CREATE EXCEL ----------------
# wb = Workbook()
# ws = wb.active
# ws.title = "Bad Images"

# headers = [
#     "Image",
#     "location",
#     "View Image",    # Hyperlink column
#     "rating_value",
#     "rating_value_ai",
#     "rating_my",      
#     "image_id",
#     "inspection_id",
#     "staff_id",
#     "updated_by"
# ]

# ws.append(headers)

# row_num = 2

# for index, row in bad_df.iterrows():

#     image_path = os.path.join(bad_images_folder, row["image_id"])
#     absolute_path = os.path.abspath(image_path)

#     # -------- Insert text data --------
#     ws.cell(row=row_num, column=2, value=row["location"])  # location near image
#     ws.cell(row=row_num, column=4, value=row["image_id"])
#     ws.cell(row=row_num, column=5, value=row["inspection_id"])
#     ws.cell(row=row_num, column=6, value=row["rating_value"])
#     ws.cell(row=row_num, column=7, value=row["rating_value_ai"])
#     ws.cell(row=row_num, column=8, value="")  # rating_my (editable)
#     ws.cell(row=row_num, column=9, value=row["staff_id"])
#     ws.cell(row=row_num, column=10, value=row["updated_by_details"])

#     # -------- Insert Image Preview --------
#     if os.path.exists(image_path):
#         img = XLImage(image_path)
#         img.width = 140    # Increased size
#         img.height = 140
#         ws.add_image(img, f"A{row_num}")
#         ws.row_dimensions[row_num].height = 110

#         # -------- Add Hyperlink --------
#         link_cell = ws.cell(row=row_num, column=3)
#         link_cell.value = "Open Image"
#         link_cell.hyperlink = absolute_path
#         link_cell.style = "Hyperlink"

#     row_num += 1

# # Adjust column widths
# column_widths = [20, 25, 18, 20, 18, 15, 18, 15, 15, 20]

# for i, width in enumerate(column_widths, start=1):
#     ws.column_dimensions[get_column_letter(i)].width = width

# # Save Excel
# wb.save(output_excel)

# print("✅ Excel file created successfully with image preview + hyperlink!")
# print("Saved at:", output_excel)
# import pandas as pd
# import os
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image as XLImage
# from openpyxl.utils import get_column_letter

# # ---------------- FILE PATHS ----------------
# rating_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_rating_202602111506.csv"
# details_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_details_202602111506.csv"

# bad_images_folder = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/rating_image_1/bad_images"

# output_excel = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/bad_images_report.xlsx"

# # ---------------- LOAD CSV ----------------
# rating_df = pd.read_csv(rating_csv, low_memory=False)
# details_df = pd.read_csv(details_csv, low_memory=False)

# rating_df.columns = rating_df.columns.str.strip()
# details_df.columns = details_df.columns.str.strip()

# rating_df["rating_value"] = rating_df["rating_value"].astype(str).str.lower().str.strip()
# rating_df["rating_value_ai"] = rating_df["rating_value_ai"].astype(str).str.lower().str.strip()

# # ---------------- MERGE ----------------
# merged_df = pd.merge(
#     rating_df,
#     details_df,
#     on="inspection_id",
#     how="inner",
#     suffixes=("_rating", "_details")
# )

# # ---------------- FILTER BAD ----------------
# bad_df = merged_df[
#     (merged_df["rating_value"] == "poor") &
#     (merged_df["rating_value_ai"] == "bad")
# ]

# downloaded_images = os.listdir(bad_images_folder)
# bad_df = bad_df[bad_df["image_id"].isin(downloaded_images)]
# # ---------------- CREATE EXCEL ----------------
# wb = Workbook()
# ws = wb.active
# ws.title = "Bad Images"

# headers = [
#     "Image",
#     "location",
#     "View Image",
#     "rating_value",
#     "rating_value_ai",
#     "rating_my",
#     "prs_coach_no",        # ✅ NEW COLUMN
#     "staff_id",
#     "updated_by",
#     "image_id",
#     "inspection_id"
# ]

# ws.append(headers)

# row_num = 2

# for index, row in bad_df.iterrows():

#     image_path = os.path.join(bad_images_folder, row["image_id"])
#     absolute_path = os.path.abspath(image_path)

#     # -------- Insert Correct Column Mapping --------
#     ws.cell(row=row_num, column=2, value=row["location"])
#     ws.cell(row=row_num, column=4, value=row["rating_value"])
#     ws.cell(row=row_num, column=5, value=row["rating_value_ai"])
#     ws.cell(row=row_num, column=6, value="")  # rating_my editable
#     ws.cell(row=row_num, column=7, value=row.get("prs_coach_no", ""))  # safe fetch
#     ws.cell(row=row_num, column=8, value=row["staff_id"])
#     ws.cell(row=row_num, column=9, value=row["updated_by_details"])
#     ws.cell(row=row_num, column=10, value=row["image_id"])
#     ws.cell(row=row_num, column=11, value=row["inspection_id"])

#     # -------- Insert Image Preview --------
#     if os.path.exists(image_path):
#         img = XLImage(image_path)
#         img.width = 140
#         img.height = 140
#         ws.add_image(img, f"A{row_num}")
#         ws.row_dimensions[row_num].height = 110

#         # -------- Hyperlink --------
#         link_cell = ws.cell(row=row_num, column=3)
#         link_cell.value = "Open Image"
#         link_cell.hyperlink = absolute_path
#         link_cell.style = "Hyperlink"

#     row_num += 1

# # Adjust column widths
# column_widths = [20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]

# for i, width in enumerate(column_widths, start=1):
#     ws.column_dimensions[get_column_letter(i)].width = width

# # Save Excel
# wb.save(output_excel)

# print("✅ Excel file created successfully with pr_coach_no column!")
# print("Saved at:", output_excel)
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------------- FILE PATHS ----------------
rating_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_rating_202602111506.csv"
details_csv = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/spark_inspection_details_202602111506.csv"

bad_images_folder = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/rating_image_1/bad_images"

output_excel = "C:/Users/Ayushi/Desktop/sanitation_yolo_project/bad_images_report.xlsx"

# ---------------- LOAD CSV ----------------
rating_df = pd.read_csv(rating_csv, low_memory=False)
details_df = pd.read_csv(details_csv, low_memory=False)

# Clean column names
rating_df.columns = rating_df.columns.str.strip()
details_df.columns = details_df.columns.str.strip()

# Normalize rating columns
rating_df["rating_value"] = rating_df["rating_value"].astype(str).str.lower().str.strip()
rating_df["rating_value_ai"] = rating_df["rating_value_ai"].astype(str).str.lower().str.strip()

# ---------------- MERGE ----------------
merged_df = pd.merge(
    rating_df,
    details_df,
    on="inspection_id",
    how="inner",
    suffixes=("_rating", "_details")
)

# ---------------- FILTER BAD ----------------
bad_df = merged_df[
    (merged_df["rating_value"] == "poor") &
    (merged_df["rating_value_ai"] == "bad")
]

# Keep only downloaded images
downloaded_images = os.listdir(bad_images_folder)
bad_df = bad_df[bad_df["image_id"].isin(downloaded_images)]

print(f"Total filtered bad records: {len(bad_df)}")

# ---------------- CREATE EXCEL ----------------
wb = Workbook()
ws = wb.active
ws.title = "Bad Images"

headers = [
    "Image",
    "Location",
    "View Image",
    "Rating Value",
    "Rating AI",
    "Rating My",
    "PRS Coach No",
    "Staff ID",
    "Updated By",
    "Image ID",
    "Inspection ID"
]

ws.append(headers)

row_num = 2

for index, row in bad_df.iterrows():

    image_path = os.path.join(bad_images_folder, row["image_id"])

    # -------- Insert Text Columns --------
    ws.cell(row=row_num, column=2, value=row.get("location", ""))
    ws.cell(row=row_num, column=4, value=row.get("rating_value", ""))
    ws.cell(row=row_num, column=5, value=row.get("rating_value_ai", ""))
    ws.cell(row=row_num, column=6, value="")  # editable column
    ws.cell(row=row_num, column=7, value=row.get("prs_coach_no", ""))
    ws.cell(row=row_num, column=8, value=row.get("staff_id", ""))
    ws.cell(row=row_num, column=9, value=row.get("updated_by_details", ""))
    ws.cell(row=row_num, column=10, value=row.get("image_id", ""))
    ws.cell(row=row_num, column=11, value=row.get("inspection_id", ""))

    # -------- Insert Image Preview --------
    if os.path.exists(image_path):
        try:
            img = XLImage(image_path)
            img.width = 140
            img.height = 140
            ws.add_image(img, f"A{row_num}")
            ws.row_dimensions[row_num].height = 110

            # -------- Local Hyperlink --------
            link_cell = ws.cell(row=row_num, column=3)
            link_cell.value = "Open Image"
            link_cell.hyperlink = image_path
            link_cell.style = "Hyperlink"

        except Exception as e:
            print(f"Error loading image {image_path}: {e}")

    row_num += 1

# Adjust column widths
column_widths = [22, 20, 18, 15, 15, 15, 18, 15, 18, 20, 20]

for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Save Excel
wb.save(output_excel)

print("✅ Excel file created successfully!")
print("Saved at:", output_excel)
